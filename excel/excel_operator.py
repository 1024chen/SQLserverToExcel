import logging
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook
from config import *


# 获取excel
def read_excel(file_path: str) -> Workbook:
    return load_workbook(file_path)


# 数据保存为excel
def data_to_excel(data: list[dict], file_path) -> bool:
    excel = Workbook()
    sheet = excel.active
    sheet.title = sheet_name
    if len(data) == 0:
        return False
    for index, column_name in enumerate(data[0].keys()):
        sheet.cell(1, index + 1, column_name)
    for row, row_data in enumerate(data):
        for col, cell in enumerate(row_data.values()):
            sheet.cell(row + 2, col + 1, cell)
    return save_excel(file_path, excel)


# 保存excel到指定路径
def save_excel(file_path: str, work_book: Workbook) -> bool:
    work_book.save(file_path)
    work_book.close()
    return os.path.exists(file_path)


# 打印结果
def res_output(res: bool):
    if res:
        logging.info(success_info)
    else:
        logging.error(fail_info)
